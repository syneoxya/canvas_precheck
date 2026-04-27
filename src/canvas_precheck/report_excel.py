# src/canvas_precheck/report_excel.py
from __future__ import annotations

import math
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _safe_str(x: Any) -> str:
    return "" if x is None else str(x)


def _find_llm_item(llm: Any, contains: str | list[str]) -> str:
    """
    Try to pull a specific section from llm.items by matching rubric_item text.
    """
    if llm is None:
        return ""
    items = getattr(llm, "items", None)
    if not items:
        return ""
    needles = [contains.lower()] if isinstance(contains, str) else [x.lower() for x in contains]
    for it in items:
        rubric = _safe_str(getattr(it, "rubric_item", "")).lower()
        if any(needle in rubric for needle in needles):
            finding = _safe_str(getattr(it, "finding", "")).strip()
            suggestion = _safe_str(getattr(it, "suggestion", "")).strip()
            if finding and suggestion:
                return f"{finding}\n\nSuggestion: {suggestion}"
            return finding or suggestion
    return ""


def _llm_scores(llm: Any) -> list[float]:
    scores = []
    for item in (getattr(llm, "items", None) or []):
        score = getattr(item, "score", None)
        if score is None:
            continue
        try:
            value = float(score)
        except (TypeError, ValueError):
            continue
        if 0 <= value <= 1:
            value *= 100
        elif 0 <= value <= 10:
            value *= 10
        scores.append(max(0, min(100, value)))
    return scores


def _normalize_score(score: Any) -> Optional[float]:
    try:
        value = float(score)
    except (TypeError, ValueError):
        return None
    if 0 <= value <= 1:
        value *= 100
    elif 0 <= value <= 10:
        value *= 10
    return max(0, min(100, value))


def _section_scores(llm: Any) -> dict[str, float]:
    raw = getattr(llm, "section_scores", None) or {}
    if not isinstance(raw, dict):
        return {}
    scores = {}
    for section, score in raw.items():
        value = _normalize_score(score)
        if value is not None:
            scores[str(section)] = value
    return scores


def _overall_from_llm(llm: Any) -> str:
    overall = _safe_str(getattr(llm, "overall", "")).strip() if llm is not None else ""
    if overall:
        return overall
    parts = []
    for item in (getattr(llm, "items", None) or []):
        rubric = _safe_str(getattr(item, "rubric_item", "")).strip()
        finding = _safe_str(getattr(item, "finding", "")).strip()
        if rubric and finding:
            parts.append(f"{rubric}: {finding}")
    return " ".join(parts)[:600]


def compute_grade_out_of_100(fb: Any, llm: Any = None) -> int:
    """
    Content-only default grading heuristic:
    - Prefer the model's aggregate overall_score when available.
    - Otherwise average independent section_scores when available.
    - Otherwise average LLM item scores when available.
    - Otherwise start at 100 and subtract only failed content tests.
    - Do not penalize lateness, filenames, duplicates, or required-file findings here.
    """
    overall_score = _normalize_score(getattr(llm, "overall_score", None))
    if overall_score is not None:
        return int(round(overall_score))

    section_scores = list(_section_scores(llm).values())
    if section_scores:
        return int(round(sum(section_scores) / len(section_scores)))

    scores = _llm_scores(llm)
    if scores:
        return int(round(sum(scores) / len(scores)))

    grade = 100
    test_results = getattr(fb, "test_results", None) or {}
    if isinstance(test_results, dict):
        for _, r in test_results.items():
            passed = bool(r.get("passed", True))
            if not passed:
                grade -= 10

    return max(0, min(100, int(grade)))


def row_from_state(state: Dict[str, Any]) -> Dict[str, Any]:
    fb = state.get("feedback")
    llm = state.get("llm")

    name = ""
    if fb is not None and getattr(fb, "metadata", None) is not None:
        name = _safe_str(getattr(fb.metadata, "student_name", ""))

    what_well = _find_llm_item(llm, ["what was done well", "what they did well"])
    what_missing = _find_llm_item(llm, ["what is missing", "missing"])
    room_improve = _find_llm_item(llm, ["what can be improved", "room for improvement"])
    overall_feedback = _overall_from_llm(llm)

    if not overall_feedback:
        overall_feedback = "No LLM overall feedback was generated."

    grade = compute_grade_out_of_100(fb, llm)
    sections = _section_scores(llm)

    return {
        "name": name,
        "what_well": what_well,
        "what_missing": what_missing,
        "room_improve": room_improve,
        "overall": overall_feedback,
        "coding_grade": round(sections["coding"]) if "coding" in sections else "",
        "writing_grade": round(sections["writing"]) if "writing" in sections else "",
        "grade": grade,
    }


def write_assignment_summary_xlsx(rows: List[Dict[str, Any]], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Name",
        "What was done well",
        "What is missing",
        "What can be improved",
        "Overall feedback",
        "Coding grade (/100)",
        "Writing grade (/100)",
        "Grade (/100)",
    ]
    ws.append(headers)

    # Header styling
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Data rows
    for r in rows:
        ws.append([
            r.get("name", ""),
            r.get("what_well", ""),
            r.get("what_missing", ""),
            r.get("room_improve", ""),
            r.get("overall", ""),
            r.get("coding_grade", ""),
            r.get("writing_grade", ""),
            r.get("grade", ""),
        ])

    # Wrap text + top align for all cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze header row + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(8)}{ws.max_row}"

    # Column widths (tuned for feedback text)
    widths = [22, 45, 45, 45, 55, 16, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
